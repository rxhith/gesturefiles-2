import tkinter as tk
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import time
import pywhatkit
from datetime import datetime
from tkinter import messagebox
import csv
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from PIL import Image, ImageTk
from tkinter import PhotoImage
import re
import time
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook


import pyttsx3

def whatsapp(msg):
    current_datetime = datetime.now()
    current_hours = int(current_datetime.strftime("%H"))
    current_minutes = int(current_datetime.strftime("%M")) + 1
    """print("Current time:")
    print("Hours:", current_hours)
    print("Minutes:", current_minutes)"""
    pywhatkit.sendwhatmsg('+917736023344',msg, current_hours, current_minutes)


def workroot():
    window.deiconify()
    new_window.withdraw()
    #new_window2.withdraw()


averages_list=[]
data = pd.read_csv('flexupdated.csv')
features = data.iloc[:, :-2]
labels = data.iloc[:, -1]
X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=0.2, random_state=42)
classifier = RandomForestClassifier()
classifier.fit(X_train, y_train)
"""def svc(pre):
    data = pd.read_csv('flexupdated.csv')
    features = data.iloc[:, :-2]
    labels = data.iloc[:, -1]
    scaler = StandardScaler()
    X_train = scaler.fit_transform(X_train)
    X_test = scaler.transform(X_test)
    X_train, X_test, y_train, y_test = train_test_split(features, labels, test_size=0.2, random_state=42)
    model = SVC()
    model.fit(X_train, y_train)

    # Predict on the test set
    y_pred = model.predict(X_test)
    y2=model.predict(np.array(pre).reshape(1, -1))
    # Evaluate the model
    print(pre)
    accuracy = accuracy_score(y_test, y2)
    print(f"Accuracy: {accuracy}")


scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)
X_train, X_test, y_train, y_test = train_test_split(features,labels, test_size=0.2, random_state=42)

# Train the model
model = SVC()
model.fit(X_train, y_train)"""


def previous():
    window.deiconify()
    new_window.withdraw()


"""def openwindow(prediction):
    #print(f"Prediction1: {prediction}")
    if prediction == " WAVE":
        a=1
        openroot(a)
    elif prediction== " UP":
        a=2
        openroot(2)
    elif prediction== " HANDCLOSE":
        a=3

        openroot(3)
    elif prediction== " THUMBS UP":
        a=4
        openroot(4)"""

def openwindow(prediction):
    #print(f"Prediction1: {prediction}")
    if prediction == "index wave":
        a=1
        print("index wave")
        print("I need FOOD!!!")
        openroot(a)
    elif prediction== "hand close":
        a=2
        print("hand close")
        print("washroom")
        openroot(2)
    elif prediction== "come":
        a=3
        print("come")
        print("give me medicine")

        openroot(3)
    elif prediction== "two fingers":
        a=4
        print("two fingers")
        print("i am okay")
        openroot(4)
    elif prediction == "thumbs up":
        a = 5
        print("thumbs up")
        print("i need water")
        openroot(5)
"""
def openroot(a):
    def update_gif(frame_index):
        #global background_image
        frame = gif_frames[frame_index]
        gif_label.config(image=frame)
        new_window.after(100, update_gif, (frame_index + 1) % frames)

    def speak_text(text):
        engine.say(text)
        engine.runAndWait()

    def slider():
        nonlocal count, text
        if count >= len(txt):
            count = -1
            text = ''
            text_label.config(text=text)
        else:
            text = text + txt[count]
            text_label.config(text=text)
            if txt[count] == ' ' or count == len(txt) - 1:
                speak_text(text.strip())
                text = ''
        count = count + 1
        text_label.after(100, slider)

    if a==1:

        new_window.deiconify()
        window.withdraw()


        txt = "I NEED FOOD!!!!"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        gif_frames = []
        gif_image = Image.open("265.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==2:
        new_window.deiconify()
        window.withdraw()
        #background_image = ImageTk.PhotoImage(Image.open("20230610_143434_0000 (2).jpg"))


        txt = "TAKE ME TO THE WASHROOM!!!!!"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        #text_label.pack(pady=(200, 50))
        text_label.pack(pady=(300, 50))

        gif_frames = []
        gif_image = Image.open("oD.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==3:

        whatsapp("i need food")
        new_window.deiconify()
        window.withdraw()


        txt = "GIVE ME MEDICINE"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        gif_frames = []
        gif_image = Image.open("94aW.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==4:
        new_window.deiconify()
        window.withdraw()

        txt = "I AM OKAY"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        gif_frames = []
        gif_image = Image.open("Vg6.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==5:
        new_window.deiconify()
        window.withdraw()

        txt = "I NEED WATER!!!!"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        gif_frames = []
        gif_image = Image.open("1NC5.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)"""
def openroot(a):
    def update_gif(frame_index):
        #global background_image
        frame = gif_frames[frame_index]
        gif_label.config(image=frame)
        new_window.after(100, update_gif, (frame_index + 1) % frames)

    def speak_text(text):
        engine.say(text)
        engine.runAndWait()

    def slider():
        nonlocal count, text
        if count >= len(txt):
            count = -1
            text = ''
            text_label.config(text=text)
        else:
            text = text + txt[count]
            text_label.config(text=text)
            if txt[count] == ' ' or count == len(txt) - 1:
                speak_text(text.strip())
                text = ''
        count = count + 1
        text_label.after(100, slider)

    if a==1:

        #new_window.deiconify()
        window.withdraw()
        whatsapp("i need food")


        txt = "I NEED FOOD!!!!"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        """gif_frames = []
        gif_image = Image.open("265.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)"""

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==2:
        #new_window.deiconify()
        window.withdraw()
        whatsapp("Take me to the washroom")
        #background_image = ImageTk.PhotoImage(Image.open("20230610_143434_0000 (2).jpg"))


        txt = "TAKE ME TO THE WASHROOM!!!!!"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        #text_label.pack(pady=(200, 50))
        text_label.pack(pady=(300, 50))

        """gif_frames = []
        gif_image = Image.open("oD.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)"""

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==3:

        whatsapp("Give me medicine")
        new_window.deiconify()
        window.withdraw()


        txt = "GIVE ME MEDICINE"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        """gif_frames = []
        gif_image = Image.open("94aW.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)"""

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        update_gif(0)
    elif a==4:
        #new_window.deiconify()
        window.withdraw()
        whatsapp("I AM OKAY")
        txt = "I AM OKAY"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        """gif_frames = []
        gif_image = Image.open("cat-cute (1).gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)"""

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        #update_gif(0)
    elif a==5:
        #new_window.deiconify()
        window.withdraw()
        whatsapp("I need Water")

        txt = "I NEED WATER!!!!"
        count = 0
        text = ''

        text_label = tk.Label(new_window, text=txt, font=('Arial', 30, "bold"), fg='black', bg='white')
        text_label.pack(pady=(300, 50))

        """gif_frames = []
        gif_image = Image.open("1NC5.gif")
        frames = gif_image.n_frames

        for i in range(frames):
            gif_image.seek(i)
            frame = ImageTk.PhotoImage(gif_image)
            gif_frames.append(frame)

        gif_label = tk.Label(new_window)
        gif_label.pack(pady=0)"""

        engine = pyttsx3.init()
        engine.setProperty('rate', 300)

        slider()
        #update_gif(0)

def get_most_recent_data():

    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('credible-acre-387913-fb62a147f8b1.json', scope)
    client = gspread.authorize(credentials)
    sheet = client.open('revamp').sheet1
    def calculate_average():
        # Get the last five values in flex1 and flex2 columns
        flex1_values = sheet.col_values(1)[-5:]  # Assuming flex1 column is the first column (column 1)
        flex2_values = sheet.col_values(2)[-5:]
        flex3_values = sheet.col_values(3)[-5:]

        # Convert the values to floats and calculate the average
        flex1_average = sum(map(float, flex1_values)) / len(flex1_values)
        flex2_average = sum(map(float, flex2_values)) / len(flex2_values)
        flex3_average = sum(map(float, flex3_values)) / len(flex3_values)


        # Do something with the average values
        print("Flex1 Average:", flex1_average)
        print("Flex2 Average:", flex2_average)
        print("Flex3 Average:", flex3_average)

        if averages_list:  # Check if the list is not empty
            averages_list.clear()  # Clear the list if it's not empty
        averages_list.append((flex1_average, flex2_average, flex3_average))
        prediction2 = classifier.predict(np.array(averages_list).reshape(1, -1))
        print(f"Prediction2: {prediction2}")
        openwindow(prediction2)






    while True:
        calculate_average()
        time.sleep(15)  # Wait for 5 seconds

def flextest():
    import re
    from openpyxl import load_workbook

    def flextest():
        book = load_workbook('q.xlsx')
        sheet = book.active
        all_rows = []
        column_averages = {'A': [], 'B': [], 'C': []}
        list1 = []

        for row_num in range(4, 50):  # Loop through rows 17 to 21
            data = {}
            value_a = sheet[f'A{row_num}'].value
            if isinstance(value_a, str):
                numeric_value_a = re.findall(r'\d+', value_a)
                if numeric_value_a:
                    data['A'] = int(numeric_value_a[0])

            data['B'] = sheet[f'B{row_num}'].value
            data['C'] = sheet[f'C{row_num}'].value
            all_rows.append(data)

            # Calculate individual column averages
            for key, value in data.items():
                if isinstance(value, (int, float)):
                    column_averages[key].append(value)

        # Calculate column averages and append to list1
        for key, values in column_averages.items():
            column_avg = sum(values) / len(values)
            list1.append(column_avg)

        # print(all_rows)
        # print('---------------------')
        print(list1)
        print('---------------------')
        prediction2 = classifier.predict(np.array(list1).reshape(1, -1))
        #svc(list1)
        print(f"Prediction2: {prediction2}")
        openwindow(prediction2)


    while True:
        flextest()
        time.sleep(10)

window = tk.Tk()
window.geometry("1450x975")

window.title("Project")

"""button = tk.Button(window, text="Predict", command=get_most_recent_data)
button.pack()"""
button_width = 40
button_height = 5
background_image2 = ImageTk.PhotoImage(Image.open("IMG-20230611-WA0108-02 (2).jpeg"))

background_label2 = tk.Label(window, image=background_image2)
background_label2.place(x=0, y=0, relwidth=1, relheight=1)
#button_x = (window.winfo_width() - button_width) // 2  # Center horizontally
#button_y = (window.winfo_height() - button_height) // 2  # Center vertically
image_c1=tk.PhotoImage(file="c82cc37b17553ca8ee5f96c7a91d53a99a3bb738_2000x2000-02 (1).png")

#button = tk.Button(window, text="Predict", command=get_most_recent_data, width=button_width, height=button_height)
button = tk.Button(window,image=image_c1, command=flextest, width=300, height=40)
button.place(x=825, y=370)


new_window = tk.Toplevel(window)
#new_window.geometry("1450x975")
new_window.geometry("1000x975")
new_window.title("come here")
background_image = ImageTk.PhotoImage(Image.open("cool futuristic 0-02 (1) (2).jpeg"))
background_label = tk.Label(new_window, image=background_image)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

"""new_window2 = tk.Toplevel(window)
new_window2.geometry("1000x975")
new_window2.title("sub")"""
workroot()
window.mainloop()
